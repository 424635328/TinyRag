from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import yaml
from langchain_community.document_loaders import PyPDFLoader, TextLoader
from langchain_community.vectorstores import FAISS
from langchain_core.documents import Document
from langchain_core.output_parsers import StrOutputParser
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_ollama import ChatOllama
from langchain_openai import ChatOpenAI
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook

from .config import RagConfig
from .prompting import build_prompt


SUPPORTED_KNOWLEDGE_SUFFIXES = (
    ".txt",
    ".md",
    ".markdown",
    ".pdf",
    ".csv",
    ".tsv",
    ".json",
    ".jsonl",
    ".yaml",
    ".yml",
    ".xlsx",
    ".xlsm",
    ".xltx",
    ".xltm",
)

QUESTION_MARKERS = (
    "?",
    "？",
    "吗",
    "么",
    "什么",
    "多少",
    "几",
    "谁",
    "哪",
    "哪里",
    "哪儿",
    "如何",
    "怎么样",
    "怎",
    "是否",
)

PRONOUN_MARKERS = ("他", "她", "它", "其", "这个", "那个", "该", "上述", "前者", "后者")

GENERIC_TERMS = {
    "rank",
    "meanrank",
    "score",
    "value",
    "field",
    "gl",
    "fata",
    "多少",
    "什么",
    "哪个",
    "哪些",
    "谁",
    "吗",
    "么",
    "如何",
    "怎么样",
}

FIELD_HINTS = {
    "rank",
    "meanrank",
    "score",
    "value",
    "accuracy",
    "acc",
    "precision",
    "recall",
    "f1",
    "auc",
    "name",
    "id",
    "编号",
    "名称",
    "分数",
    "排名",
    "均值",
    "平均",
}

IDENTIFIER_FIELD_HINTS = {
    "name",
    "id",
    "key",
    "algorithm",
    "model",
    "method",
    "object",
    "entity",
    "名称",
    "编号",
    "对象",
    "实体",
    "算法",
    "模型",
    "方法",
}


@dataclass
class PreparedQuery:
    original_question: str
    rewritten_question: str
    history_text: str
    selected_docs: List[Document]
    evidence: List[dict]
    entities: List[str]
    debug: dict
    direct_answer: Optional[str] = None


def _resolve_device(preferred: str) -> str:
    if preferred != "auto":
        return preferred

    try:
        import torch

        return "cuda" if torch.cuda.is_available() else "cpu"
    except Exception:
        return "cpu"


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", "", text.strip().lower())


def _stringify_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _safe_header(value, fallback: str) -> str:
    text = _stringify_value(value)
    return text or fallback


def _build_row_text(items) -> str:
    pairs = []
    for key, value in items:
        text = _stringify_value(value)
        if not text:
            continue
        pairs.append("{0}: {1}".format(key, text))
    return "；".join(pairs)


def _build_structured_metadata(source: str, record_index: int, record: dict, base_metadata=None):
    metadata = dict(base_metadata or {})
    structured_record = {}
    structured_values = []
    for key, value in record.items():
        text = _stringify_value(value)
        structured_record[str(key)] = text
        if text:
            structured_values.append(text)

    metadata.update(
        {
            "source": source,
            "record_index": record_index,
            "structured_fields": list(structured_record.keys()),
            "structured_values": structured_values,
            "structured_record": structured_record,
        }
    )
    return metadata


def _records_to_documents(records, source: str, base_metadata=None):
    documents = []
    for index, record in enumerate(records, start=1):
        text = _build_row_text(record.items())
        if not text:
            continue
        metadata = _build_structured_metadata(source, index, record, base_metadata)
        documents.append(Document(page_content=text, metadata=metadata))
    return documents


def _load_delimited_file(path: Path, delimiter: str):
    source = str(path).replace("\\", "/")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return []

    headers = [_safe_header(value, "列{0}".format(index + 1)) for index, value in enumerate(rows[0])]
    data_rows = rows[1:] if any(header.strip() for header in headers) else rows
    if data_rows is rows:
        headers = ["列{0}".format(index + 1) for index in range(len(rows[0]))]

    records = []
    for row in data_rows:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        record = dict(zip(headers, padded[: len(headers)]))
        records.append(record)

    return _records_to_documents(records, source, {"file_type": path.suffix.lower()})


def _normalize_object_for_documents(data):
    if isinstance(data, list):
        records = []
        for item in data:
            if isinstance(item, dict):
                records.append(item)
            else:
                records.append({"value": item})
        return records

    if isinstance(data, dict):
        if all(isinstance(value, dict) for value in data.values()):
            records = []
            for key, value in data.items():
                record = {"key": key}
                record.update(value)
                records.append(record)
            return records
        return [data]

    return [{"value": data}]


def _load_json_file(path: Path):
    source = str(path).replace("\\", "/")
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    records = _normalize_object_for_documents(data)
    return _records_to_documents(records, source, {"file_type": path.suffix.lower()})


def _load_jsonl_file(path: Path):
    source = str(path).replace("\\", "/")
    records = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            item = json.loads(line)
            if isinstance(item, dict):
                records.append(item)
            else:
                records.append({"value": item})
    return _records_to_documents(records, source, {"file_type": path.suffix.lower()})


def _load_yaml_file(path: Path):
    source = str(path).replace("\\", "/")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    records = _normalize_object_for_documents(data)
    return _records_to_documents(records, source, {"file_type": path.suffix.lower()})


def _load_excel_file(path: Path):
    source = str(path).replace("\\", "/")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    documents = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [
            _safe_header(value, "列{0}".format(index + 1))
            for index, value in enumerate(rows[0])
        ]
        data_rows = rows[1:] if any(header.strip() for header in headers) else rows
        if data_rows is rows:
            headers = ["列{0}".format(index + 1) for index in range(len(rows[0]))]

        records = []
        for row in data_rows:
            padded = list(row) + [None] * max(0, len(headers) - len(row))
            record = dict(zip(headers, padded[: len(headers)]))
            records.append(record)

        documents.extend(
            _records_to_documents(
                records,
                source,
                {
                    "file_type": path.suffix.lower(),
                    "sheet_name": sheet.title,
                },
            )
        )

    workbook.close()
    return documents


def _load_single_file(path: Path):
    suffix = path.suffix.lower()
    source = str(path).replace("\\", "/")

    if suffix == ".pdf":
        docs = PyPDFLoader(str(path)).load()
        for doc in docs:
            doc.metadata["file_type"] = suffix
    elif suffix in (".txt", ".md", ".markdown"):
        docs = TextLoader(str(path), encoding="utf-8").load()
        for doc in docs:
            doc.metadata["file_type"] = suffix
    elif suffix == ".csv":
        docs = _load_delimited_file(path, ",")
    elif suffix == ".tsv":
        docs = _load_delimited_file(path, "\t")
    elif suffix == ".json":
        docs = _load_json_file(path)
    elif suffix == ".jsonl":
        docs = _load_jsonl_file(path)
    elif suffix in (".yaml", ".yml"):
        docs = _load_yaml_file(path)
    elif suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        docs = _load_excel_file(path)
    else:
        raise ValueError("暂不支持的知识库文件类型: {0}".format(path.suffix))

    for doc in docs:
        doc.metadata["source"] = source

    return docs


def list_knowledge_files(knowledge_path: Path) -> List[Path]:
    if not knowledge_path.exists():
        raise FileNotFoundError("知识库路径不存在: {0}".format(knowledge_path))

    if knowledge_path.is_file():
        files = [knowledge_path]
    else:
        files = [
            path
            for path in knowledge_path.rglob("*")
            if path.is_file() and path.suffix.lower() in SUPPORTED_KNOWLEDGE_SUFFIXES
        ]

    supported = [path for path in files if path.suffix.lower() in SUPPORTED_KNOWLEDGE_SUFFIXES]
    if not supported:
        raise FileNotFoundError(
            "知识库中未找到受支持文件，当前支持: {0}".format(
                ", ".join(SUPPORTED_KNOWLEDGE_SUFFIXES)
            )
        )

    return sorted(supported, key=lambda path: str(path).lower())


def describe_knowledge_snapshot(knowledge_path: Path) -> Tuple[Tuple[str, int, int], ...]:
    snapshot = []
    for path in list_knowledge_files(knowledge_path):
        stat = path.stat()
        snapshot.append((str(path.resolve()), stat.st_mtime_ns, stat.st_size))
    return tuple(snapshot)


def load_knowledge_documents(knowledge_path: Path):
    documents = []
    for path in list_knowledge_files(knowledge_path):
        documents.extend(_load_single_file(path))
    return documents


def _extract_search_terms(text: str) -> List[str]:
    lowered = text.strip().lower()
    terms = []
    seen = set()

    def add_term(term: str):
        cleaned = term.strip().lower()
        if len(cleaned) < 2 or cleaned in seen:
            return
        seen.add(cleaned)
        terms.append(cleaned)

    add_term(lowered)

    for token in re.findall(r"[a-zA-Z][a-zA-Z0-9._-]*", lowered):
        add_term(token)

    for chunk in re.findall(r"[\u4e00-\u9fff]{2,}", lowered):
        add_term(chunk)
        simplified = chunk
        for marker in QUESTION_MARKERS:
            simplified = simplified.replace(marker, " ")
        for part in simplified.split():
            add_term(part)

    return terms


def _extract_entities(text: str, known_fields=None) -> List[str]:
    known_fields = {field.lower() for field in (known_fields or [])}
    entities = []
    seen = set()

    def add_entity(value: str):
        cleaned = value.strip()
        lowered = cleaned.lower()
        if len(cleaned) < 2 or lowered in seen or lowered in GENERIC_TERMS or lowered in known_fields:
            return
        seen.add(lowered)
        entities.append(cleaned)

    for match in re.findall(r"([A-Za-z][A-Za-z0-9._-]{1,})", text):
        add_entity(match)

    for match in re.findall(r"([\u4e00-\u9fffA-Za-z0-9._-]{2,30})的", text):
        add_entity(match)

    if len(text.strip()) <= 24 and not any(marker in text for marker in QUESTION_MARKERS):
        add_entity(text.strip())

    return entities


def _split_question_focus(question: str) -> Tuple[Optional[str], str]:
    query = question.strip().strip("。？！?!. ")
    if not query or "的" not in query:
        return None, query

    entity_part, field_part = query.split("的", 1)
    entity = entity_part.strip().strip("“”\"' ")
    field_query = field_part.strip()
    if not entity:
        return None, query
    return entity, field_query or query


def _extract_question_entity(question: str) -> Optional[str]:
    entity, _ = _split_question_focus(question)
    if entity and entity not in PRONOUN_MARKERS:
        return entity

    entities = _extract_entities(question)
    for item in entities:
        if item not in PRONOUN_MARKERS:
            return item
    return None


def _resolve_recent_entity(history_turns, recent_entities: Sequence[str]) -> str:
    for entity in recent_entities:
        if entity and entity not in PRONOUN_MARKERS:
            return entity

    for turn in reversed(list(history_turns or [])):
        for entity in turn.get("entities", []):
            if entity and entity not in PRONOUN_MARKERS:
                return entity
        candidate = _extract_question_entity(
            turn.get("rewritten_question") or turn.get("user_question", "")
        )
        if candidate:
            return candidate
    return ""


def _looks_like_topic_query(question: str) -> bool:
    query = question.strip()
    if not query:
        return False
    if any(marker in query for marker in QUESTION_MARKERS):
        return False
    return len(query) <= 24


def _needs_context_resolution(question: str) -> bool:
    query = question.strip()
    if not query:
        return False
    if any(marker in query for marker in PRONOUN_MARKERS):
        return True

    query_terms = _extract_search_terms(query)
    explicit_entities = _extract_entities(query)
    field_like = any(term.lower() in FIELD_HINTS or "rank" in term.lower() for term in query_terms)
    return field_like and not explicit_entities


def _prefix_entity(question: str, entity: str) -> str:
    query = question.strip()
    if not query:
        return entity
    if query.startswith(entity):
        return query
    if query.startswith("的"):
        query = query[1:]
    return "{0}的{1}".format(entity, query)


def _rewrite_question(question: str, history_turns, recent_entities: Sequence[str]):
    original = question.strip()
    recent_entity = _resolve_recent_entity(history_turns, recent_entities)
    rewritten = original
    reason = "none"

    if recent_entity and any(marker in original for marker in PRONOUN_MARKERS):
        for marker in PRONOUN_MARKERS:
            rewritten = rewritten.replace(marker, recent_entity)
        reason = "pronoun_resolution"
    elif recent_entity and _needs_context_resolution(original):
        rewritten = _prefix_entity(original, recent_entity)
        reason = "entity_prefix"

    if rewritten == original and _looks_like_topic_query(original):
        reason = "topic_summary"

    return rewritten, reason


def _format_history(history_turns) -> str:
    if not history_turns:
        return "无"

    lines = []
    for turn in history_turns[-4:]:
        lines.append("用户: {0}".format(turn.get("user_question", "")))
        rewritten = turn.get("rewritten_question") or ""
        if rewritten and rewritten != turn.get("user_question", ""):
            lines.append("系统理解为: {0}".format(rewritten))
        lines.append("助手: {0}".format(turn.get("answer", "")))

    return "\n".join(lines)


def _doc_signature(doc: Document):
    return (
        doc.metadata.get("source", "unknown"),
        doc.metadata.get("sheet_name"),
        doc.metadata.get("page"),
        doc.metadata.get("record_index"),
        doc.page_content,
    )


def _score_text_match(terms: Sequence[str], text: str) -> int:
    score = 0
    lowered = text.lower()
    compact = _normalize_text(text)
    for term in terms:
        if term in lowered:
            score += len(term) * 4
        if term in compact:
            score += len(term) * 2
    return score


def _keyword_match_documents(queries: Sequence[str], documents, limit: int):
    scored = {}
    doc_map = {}
    for query in queries:
        terms = _extract_search_terms(query)
        if not terms:
            continue
        for doc in documents:
            score = _score_text_match(terms, doc.page_content)
            if score <= 0:
                continue
            signature = _doc_signature(doc)
            scored[signature] = max(scored.get(signature, 0), score)
            doc_map[signature] = doc

    ranked = sorted(scored.items(), key=lambda item: item[1], reverse=True)
    return [(score, doc_map[signature], "keyword") for signature, score in ranked[:limit]]


def _field_name_score(term: str, field_name: str) -> int:
    normalized_term = _normalize_text(term)
    normalized_field = _normalize_text(field_name)
    if not normalized_term or not normalized_field:
        return 0
    if normalized_term == normalized_field:
        return len(normalized_term) * 12
    if normalized_field.startswith(normalized_term):
        return len(normalized_term) * 9
    if normalized_term in normalized_field:
        return len(normalized_term) * 7
    return 0


def _field_match_documents(queries: Sequence[str], documents, limit: int):
    scored = {}
    doc_map = {}
    for query in queries:
        terms = _extract_search_terms(query)
        if not terms:
            continue
        for doc in documents:
            fields = doc.metadata.get("structured_fields", [])
            values = doc.metadata.get("structured_values", [])
            sheet_name = doc.metadata.get("sheet_name", "")
            score = 0
            for term in terms:
                score += sum(_field_name_score(term, field) for field in fields)
                score += _score_text_match([term], " ".join(values))
                if sheet_name:
                    score += _score_text_match([term], str(sheet_name))
            if score <= 0:
                continue
            signature = _doc_signature(doc)
            scored[signature] = max(scored.get(signature, 0), score)
            doc_map[signature] = doc

    ranked = sorted(scored.items(), key=lambda item: item[1], reverse=True)
    return [(score, doc_map[signature], "field") for signature, score in ranked[:limit]]


def _semantic_match_documents(queries: Sequence[str], retriever, limit: int):
    scored = {}
    doc_map = {}
    for query_index, query in enumerate(queries):
        docs = retriever.invoke(query)
        for rank, doc in enumerate(docs):
            score = max((limit - rank), 1) * (8 - query_index)
            signature = _doc_signature(doc)
            scored[signature] = max(scored.get(signature, 0), score)
            doc_map[signature] = doc

    ranked = sorted(scored.items(), key=lambda item: item[1], reverse=True)
    return [(score, doc_map[signature], "semantic") for signature, score in ranked[:limit]]


def _filter_low_score_threshold() -> int:
    return 12


def _is_text_document(doc: Document) -> bool:
    file_type = doc.metadata.get("file_type", "")
    return file_type in (".txt", ".md", ".markdown", ".pdf")


def _is_structured_document(doc: Document) -> bool:
    file_type = doc.metadata.get("file_type", "")
    return file_type in (".csv", ".tsv", ".json", ".jsonl", ".yaml", ".yml", ".xlsx", ".xlsm", ".xltx", ".xltm")


def _has_keyword_match(terms: Sequence[str], text: str) -> bool:
    if not terms:
        return False
    lowered = text.lower()
    for term in terms:
        if term in lowered:
            return True
    return False


def _is_doc_relevant(terms: Sequence[str], doc: Document) -> bool:
    if not terms:
        return True
    
    text = doc.page_content.lower()
    
    for term in terms:
        if term in text:
            return True
    
    if _is_structured_document(doc):
        structured_values = doc.metadata.get("structured_values", [])
        for value in structured_values:
            value_lower = str(value).lower()
            for term in terms:
                if term in value_lower:
                    return True
    
    return False


def _should_include_structured_doc(terms: Sequence[str], doc: Document, score: int) -> bool:
    if not _is_structured_document(doc):
        return True
    
    if score < 20:
        return False
    
    if _has_keyword_match(terms, doc.page_content):
        return True
    
    structured_values = doc.metadata.get("structured_values", [])
    for value in structured_values:
        if _has_keyword_match(terms, str(value)):
            return True
    
    return False


def _merge_scored_documents(*groups):
    merged = {}
    for group in groups:
        for score, doc, reason in group:
            signature = _doc_signature(doc)
            if signature not in merged:
                merged[signature] = {"doc": doc, "score": 0, "reasons": []}
            merged[signature]["score"] += score
            if reason not in merged[signature]["reasons"]:
                merged[signature]["reasons"].append(reason)

    ranked = sorted(
        merged.values(),
        key=lambda item: (item["score"], len(item["doc"].page_content)),
        reverse=True,
    )
    return ranked


def _doc_preview(doc: Document, max_length: int = 180) -> str:
    text = re.sub(r"\s+", " ", doc.page_content).strip()
    if len(text) <= max_length:
        return text
    return text[: max_length - 1] + "…"


def _evidence_item(doc: Document, score: int, reasons: Sequence[str]) -> dict:
    metadata = doc.metadata
    return {
        "source": metadata.get("source", "unknown"),
        "file_type": metadata.get("file_type"),
        "sheet_name": metadata.get("sheet_name"),
        "page": metadata.get("page"),
        "record_index": metadata.get("record_index"),
        "score": score,
        "reasons": list(reasons),
        "preview": _doc_preview(doc),
        "fields": metadata.get("structured_fields", [])[:8],
    }


def _answer_indicates_no_info(answer: str) -> bool:
    text = answer.strip()
    return "未提供相关信息" in text or "资料未提供" in text


def _extract_referenced_entity(rewritten_question: str, selected_docs: Sequence[Document]) -> Optional[str]:
    explicit_entity = _extract_question_entity(rewritten_question)
    if explicit_entity:
        return explicit_entity

    fields = []
    for doc in selected_docs:
        fields.extend(doc.metadata.get("structured_fields", []))
    entities = _extract_entities(rewritten_question, fields)
    if entities:
        return entities[0]
    return None


def _is_identifier_field(field_name: str) -> bool:
    normalized = _normalize_text(field_name)
    if not normalized:
        return False
    if normalized in IDENTIFIER_FIELD_HINTS:
        return True
    return any(hint in normalized for hint in ("name", "algorithm", "model", "method"))


def _record_entity_score(entity: Optional[str], record: dict) -> int:
    if not entity:
        return 0

    normalized_entity = entity.lower()
    best_score = 0
    for field_name, value in record.items():
        text = _stringify_value(value)
        if not text:
            continue

        score = _score_text_match([normalized_entity], text)
        if _normalize_text(text) == _normalize_text(entity):
            score += max(len(entity), 2) * 20
        elif _normalize_text(entity) in _normalize_text(text):
            score += max(len(entity), 2) * 8

        if score and _is_identifier_field(field_name):
            score += max(len(entity), 2) * 10

        best_score = max(best_score, score)

    return best_score


def _select_structured_field(question: str, docs: Sequence[Document]) -> Tuple[Optional[Document], Optional[str], Optional[str]]:
    best = None
    referenced_entity, field_query = _split_question_focus(question)
    if referenced_entity in PRONOUN_MARKERS:
        referenced_entity = None
    if not referenced_entity:
        referenced_entity = _extract_referenced_entity(question, docs)

    question_terms = _extract_search_terms(field_query or question)

    for doc in docs:
        record = doc.metadata.get("structured_record")
        if not record:
            continue

        entity_score = _record_entity_score(referenced_entity, record)

        best_field = None
        best_field_score = 0
        for field_name, value in record.items():
            if not value:
                continue
            score = 0
            for term in question_terms:
                score += _field_name_score(term, field_name)
            if score > best_field_score:
                best_field_score = score
                best_field = field_name

        if referenced_entity and entity_score <= 0:
            continue
        if best_field_score <= 0:
            continue

        total_score = entity_score + best_field_score
        if best_field and total_score > 0:
            candidate = (total_score, doc, best_field, record.get(best_field, ""))
            if best is None or candidate[0] > best[0]:
                best = candidate

    if not best:
        return None, None, None
    return best[1], best[2], best[3]


def _direct_structured_answer(question: str, docs: Sequence[Document]) -> Optional[str]:
    doc, field_name, value = _select_structured_field(question, docs)
    if not doc or not field_name or not value:
        return None

    entity = _extract_question_entity(question) or _extract_referenced_entity(question, [doc]) or "该对象"
    return "根据参考资料，{0}的{1}是{2}。".format(entity, field_name, value)


def _normalize_question_for_generation(question: str, matched_docs, rewritten_question: str) -> str:
    direct_answer = _direct_structured_answer(rewritten_question, matched_docs)
    if direct_answer:
        return rewritten_question

    query = rewritten_question.strip()
    if _looks_like_topic_query(query) and matched_docs:
        return (
            "请根据参考资料，概括与“{0}”直接相关的已知信息。"
            "如果资料中提到偏好、身份、定义、制度或事实，请直接总结。"
        ).format(query)
    return query


def _build_llm(config: RagConfig):
    if config.llm_provider == "ollama":
        kwargs = {
            "model": config.ollama_model,
            "temperature": config.temperature,
        }
        if config.ollama_base_url:
            kwargs["base_url"] = config.ollama_base_url
        return ChatOllama(**kwargs)

    if config.llm_provider == "deepseek":
        if not config.deepseek_api_key:
            raise ValueError("使用 DeepSeek 时必须提供 DEEPSEEK_API_KEY。")
        return ChatOpenAI(
            model=config.deepseek_model,
            api_key=config.deepseek_api_key,
            base_url=config.deepseek_base_url,
            temperature=config.temperature,
        )

    raise ValueError("不支持的 LLM_PROVIDER: {0}".format(config.llm_provider))


def _extract_chunk_text(chunk) -> str:
    content = getattr(chunk, "content", chunk)
    if isinstance(content, list):
        texts = []
        for item in content:
            if isinstance(item, dict):
                texts.append(str(item.get("text", "")))
            else:
                texts.append(str(item))
        return "".join(texts)
    return str(content or "")


class RagEngine:
    def __init__(self, config: RagConfig) -> None:
        self.config = config
        self.raw_documents = load_knowledge_documents(config.knowledge_path)
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=config.chunk_size,
            chunk_overlap=config.chunk_overlap,
            separators=["\n\n", "\n", "。", "！", "？", "，", "、", " ", ""],
        )
        self.chunks = self.splitter.split_documents(self.raw_documents)
        self.embeddings = HuggingFaceEmbeddings(
            model_name=config.embedding_model_name,
            model_kwargs={"device": _resolve_device(config.embedding_device)},
            encode_kwargs={"normalize_embeddings": True},
        )
        self.vectorstore = FAISS.from_documents(documents=self.chunks, embedding=self.embeddings)
        self.semantic_limit = max(config.retriever_k * 3, 6)
        self.final_context_limit = max(config.retriever_k * 3, 6)
        self.retriever = self.vectorstore.as_retriever(
            search_type="similarity",
            search_kwargs={"k": self.semantic_limit},
        )
        self.prompt = build_prompt()
        self.llm = _build_llm(config)
        self.output_parser = StrOutputParser()

    def prepare(self, question: str, history_turns=None, recent_entities=None) -> PreparedQuery:
        history_turns = list(history_turns or [])
        recent_entities = list(recent_entities or [])
        original_question = question.strip()
        rewritten_question, rewrite_reason = _rewrite_question(
            original_question,
            history_turns,
            recent_entities,
        )

        retrieval_queries = [rewritten_question]
        if rewritten_question != original_question:
            retrieval_queries.append(original_question)
        if history_turns and _needs_context_resolution(original_question):
            last_turn = history_turns[-1]
            last_question = last_turn.get("rewritten_question") or last_turn.get("user_question", "")
            if last_question and last_question not in retrieval_queries:
                retrieval_queries.append(last_question)

        semantic_docs = _semantic_match_documents(retrieval_queries, self.retriever, self.semantic_limit)
        keyword_docs = _keyword_match_documents(retrieval_queries, self.chunks, self.semantic_limit)
        field_docs = _field_match_documents(retrieval_queries, self.chunks, self.semantic_limit)
        ranked_docs = _merge_scored_documents(semantic_docs, keyword_docs, field_docs)

        search_terms = _extract_search_terms(rewritten_question)
        min_score_threshold = _filter_low_score_threshold()
        
        filtered_ranked_docs = []
        for item in ranked_docs:
            if item["score"] < min_score_threshold:
                continue
            
            if not _is_doc_relevant(search_terms, item["doc"]):
                continue
            
            if not _should_include_structured_doc(search_terms, item["doc"], item["score"]):
                continue
            
            filtered_ranked_docs.append(item)
        
        if not filtered_ranked_docs and ranked_docs:
            for item in ranked_docs:
                if _is_text_document(item["doc"]) and _is_doc_relevant(search_terms, item["doc"]):
                    filtered_ranked_docs = [item]
                    break
            
            if not filtered_ranked_docs:
                filtered_ranked_docs = ranked_docs[:1]

        selected_ranked_docs = filtered_ranked_docs[: self.final_context_limit]
        selected_docs = [item["doc"] for item in selected_ranked_docs]
        evidence = [
            _evidence_item(item["doc"], item["score"], item["reasons"])
            for item in selected_ranked_docs[: min(6, len(selected_ranked_docs))]
        ]
        generation_question = _normalize_question_for_generation(
            original_question,
            selected_docs,
            rewritten_question,
        )
        combined_entities = []
        explicit_entity = _extract_question_entity(rewritten_question) or _extract_question_entity(
            original_question
        )
        if explicit_entity:
            combined_entities.append(explicit_entity)

        for entity in _extract_entities(
            rewritten_question,
            [field for doc in selected_docs for field in doc.metadata.get("structured_fields", [])],
        ):
            if entity not in combined_entities:
                combined_entities.append(entity)
        if not combined_entities:
            combined_entities = recent_entities[:1]

        debug = {
            "original_question": original_question,
            "rewritten_question": generation_question,
            "rewrite_reason": rewrite_reason,
            "history_turns_used": len(history_turns[-4:]),
            "recent_entities": recent_entities,
            "retrieval_queries": retrieval_queries,
            "matched_sources": [item["source"] for item in evidence],
        }

        return PreparedQuery(
            original_question=original_question,
            rewritten_question=generation_question,
            history_text=_format_history(history_turns),
            selected_docs=selected_docs,
            evidence=evidence,
            entities=combined_entities,
            debug=debug,
            direct_answer=_direct_structured_answer(generation_question, selected_docs),
        )

    def _build_prompt_value(self, prepared: PreparedQuery):
        return self.prompt.invoke(
            {
                "history": prepared.history_text,
                "original_question": prepared.original_question,
                "question": prepared.rewritten_question,
                "context": self._format_context(prepared.selected_docs),
            }
        )

    def _format_context(self, docs: Sequence[Document]) -> str:
        if not docs:
            return "无"
        parts = []
        for index, doc in enumerate(docs, start=1):
            source = doc.metadata.get("source", "unknown")
            file_type = doc.metadata.get("file_type", "")
            sheet_name = doc.metadata.get("sheet_name")
            page = doc.metadata.get("page")
            record_index = doc.metadata.get("record_index")

            labels = ["片段 {0}".format(index), "来源: {0}".format(source)]
            if file_type:
                labels.append("类型: {0}".format(file_type))
            if sheet_name:
                labels.append("工作表: {0}".format(sheet_name))
            if page is not None:
                labels.append("页码: {0}".format(page))
            if record_index is not None:
                labels.append("记录: {0}".format(record_index))
            parts.append("[{0}]\n{1}".format(" | ".join(labels), doc.page_content))
        return "\n\n".join(parts)

    def generate(self, prepared: PreparedQuery) -> str:
        if prepared.direct_answer:
            return prepared.direct_answer
        message = self.llm.invoke(self._build_prompt_value(prepared))
        return self.output_parser.invoke(message)

    def stream_generate(self, prepared: PreparedQuery):
        if prepared.direct_answer:
            yield prepared.direct_answer
            return
        for chunk in self.llm.stream(self._build_prompt_value(prepared)):
            text = _extract_chunk_text(chunk)
            if text:
                yield text


def build_rag_chain(config: RagConfig):
    return RagEngine(config)
